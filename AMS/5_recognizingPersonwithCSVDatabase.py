from collections.abc import Iterable
import numpy as np
import imutils
import pickle
import time
import cv2
import csv
from openpyxl import load_workbook
from datetime import datetime


def flatten(lis):
    for item in lis:
        if isinstance(item, Iterable) and not isinstance(item, str):
            for x in flatten(item):
                yield x
        else:
            yield item


embeddingFile = "output/embeddings.pickle"
embeddingModel = "openface_nn4.small2.v1.t7"
recognizerFile = "output/recognizer.pickle"
labelEncFile = "output/le.pickle"
conf = 0.5
attendanceFile = "attendance.xlsx"  # Your Excel attendance file

print("[INFO] loading face detector...")
prototxt = "model/deploy.prototxt"
model = "model/res10_300x300_ssd_iter_140000.caffemodel"
detector = cv2.dnn.readNetFromCaffe(prototxt, model)

print("[INFO] loading face recognizer...")
embedder = cv2.dnn.readNetFromTorch(embeddingModel)

recognizer = pickle.loads(open(recognizerFile, "rb").read())
le = pickle.loads(open(labelEncFile, "rb").read())

Roll_Number = ""
box = []
print("[INFO] starting video stream...")
cam = cv2.VideoCapture(0)
time.sleep(2.0)

# Load the Excel file for attendance
def mark_attendance(name, roll_number):
    date = datetime.now().strftime('%Y-%m-%d')  # Current date
    status = 'Present'

    try:
        wb = load_workbook(attendanceFile)
        sheet = wb.active

        # Check if the person's name and date already exist
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            if row[0].value == name and row[1].value == roll_number and row[2].value == date:
                print(f"[INFO] Attendance already marked for {name} today.")
                return

        # If not, add new attendance record
        sheet.append([name, roll_number, date, status])
        wb.save(attendanceFile)
        print(f"[INFO] Attendance marked for {name} on {date}.")
    except FileNotFoundError:
        print(f"[ERROR] The attendance file {attendanceFile} does not exist.")
        return

while True:
    _, frame = cam.read()
    frame = imutils.resize(frame, width=600)
    (h, w) = frame.shape[:2]
    imageBlob = cv2.dnn.blobFromImage(
        cv2.resize(frame, (300, 300)), 1.0, (300, 300),
        (104.0, 177.0, 123.0), swapRB=False, crop=False)

    detector.setInput(imageBlob)
    detections = detector.forward()

    for i in range(0, detections.shape[2]):

        confidence = detections[0, 0, i, 2]

        if confidence > conf:

            box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
            (startX, startY, endX, endY) = box.astype("int")

            face = frame[startY:endY, startX:endX]
            (fH, fW) = face.shape[:2]

            if fW < 20 or fH < 20:
                continue

            faceBlob = cv2.dnn.blobFromImage(face, 1.0 / 255, (96, 96), (0, 0, 0), swapRB=True, crop=False)
            embedder.setInput(faceBlob)
            vec = embedder.forward()

            preds = recognizer.predict_proba(vec)[0]
            j = np.argmax(preds)
            proba = preds[j]
            name = le.classes_[j]

            # Get roll number and attendance from the CSV
            Roll_Number = ""
            with open('student.csv', 'r') as csvFile:
                reader = csv.reader(csvFile)
                for row in reader:
                    name_str = str(name)
                    if name_str in row:
                        Roll_Number = row[1]  # Assuming roll number is in the second column
                        print(f"Roll Number for {name}: {Roll_Number}")
                        break

            # Mark attendance
            if Roll_Number:
                mark_attendance(name, Roll_Number)

            # Display the result on screen
            text = "{} : {} : {:.2f}%".format(name, Roll_Number, proba * 100)
            y = startY - 10 if startY - 10 > 10 else startY + 10
            cv2.rectangle(frame, (startX, startY), (endX, endY),
                          (0, 0, 255), 2)
            cv2.putText(frame, text, (startX, y),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 255), 2)

    # Show the frame
    cv2.imshow("Frame", frame)
    key = cv2.waitKey(1) & 0xFF
    if key == 27:  # Press 'ESC' to exit
        break
   

cam.release()
cv2.destroyAllWindows()
